"""
Xuất kết quả so sánh phân loại cũ và mới ra file Excel
"""

import sys
sys.stdout.reconfigure(encoding='utf-8')

import json
import os
from datetime import datetime
from supabase import create_client
from data_generator import StudentDataGenerator
from student_classifier import StudentClassifier
from skill_evaluator import SkillEvaluator

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows
except ImportError:
    print("⚠️ Cần cài đặt openpyxl: pip install openpyxl")
    sys.exit(1)

SUPABASE_URL = "https://odmtndvllclmrwczcyvs.supabase.co"
SUPABASE_KEY = "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6Im9kbXRuZHZsbGNsbXJ3Y3pjeXZzIiwicm9sZSI6ImFub24iLCJpYXQiOjE3NjQwNDI0NDIsImV4cCI6MjA3OTYxODQ0Mn0.au4mfOQSocrCr9eC753wiveR1KI0TNAVxOk1KB5poMA"

HISTORY_FILE = 'classification_history.json'

def load_history():
    """Load lịch sử phân loại từ file JSON"""
    if os.path.exists(HISTORY_FILE):
        with open(HISTORY_FILE, 'r', encoding='utf-8') as f:
            return json.load(f)
    return {'versions': [], 'current_version': 0}

def get_current_classifications(supabase):
    """Lấy kết quả phân loại hiện tại từ Supabase"""
    result = supabase.table('classifications').select('*').execute()
    return {c['student_id']: c for c in result.data}

def create_excel_report(students_info, old_data, new_data, output_file):
    """Tạo file Excel so sánh"""
    wb = openpyxl.Workbook()
    
    # Styles
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="667eea", end_color="667eea", fill_type="solid")
    
    level_fills = {
        'Xuat sac': PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid"),
        'Kha': PatternFill(start_color="BBDEFB", end_color="BBDEFB", fill_type="solid"),
        'Trung binh': PatternFill(start_color="FFE0B2", end_color="FFE0B2", fill_type="solid"),
        'Yeu': PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")
    }
    
    change_fills = {
        'up': PatternFill(start_color="A5D6A7", end_color="A5D6A7", fill_type="solid"),
        'down': PatternFill(start_color="EF9A9A", end_color="EF9A9A", fill_type="solid"),
        'same': PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    }
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    center_align = Alignment(horizontal='center', vertical='center')
    
    # ===== SHEET 1: Tổng quan =====
    ws1 = wb.active
    ws1.title = "Tổng quan"
    
    # Header
    headers = ["Chỉ số", "Giá trị cũ", "Giá trị mới", "Thay đổi"]
    for col, header in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border
    
    # Thống kê
    old_stats = {'Xuat sac': 0, 'Kha': 0, 'Trung binh': 0, 'Yeu': 0, 'anomaly': 0}
    new_stats = {'Xuat sac': 0, 'Kha': 0, 'Trung binh': 0, 'Yeu': 0, 'anomaly': 0}
    
    for sid, data in old_data.items():
        level = data.get('final_level', '')
        if level in old_stats:
            old_stats[level] += 1
        if data.get('anomaly_detected'):
            old_stats['anomaly'] += 1
    
    for sid, data in new_data.items():
        level = data.get('final_level', '')
        if level in new_stats:
            new_stats[level] += 1
        if data.get('anomaly_detected'):
            new_stats['anomaly'] += 1
    
    stats_rows = [
        ("Tổng sinh viên", len(old_data), len(new_data)),
        ("Xuất sắc", old_stats['Xuat sac'], new_stats['Xuat sac']),
        ("Khá", old_stats['Kha'], new_stats['Kha']),
        ("Trung bình", old_stats['Trung binh'], new_stats['Trung binh']),
        ("Yếu", old_stats['Yeu'], new_stats['Yeu']),
        ("Bất thường", old_stats['anomaly'], new_stats['anomaly'])
    ]
    
    for row_idx, (label, old_val, new_val) in enumerate(stats_rows, 2):
        diff = new_val - old_val
        diff_str = f"+{diff}" if diff > 0 else str(diff)
        
        ws1.cell(row=row_idx, column=1, value=label).border = thin_border
        ws1.cell(row=row_idx, column=2, value=old_val).border = thin_border
        ws1.cell(row=row_idx, column=3, value=new_val).border = thin_border
        ws1.cell(row=row_idx, column=4, value=diff_str).border = thin_border
        
        for col in range(1, 5):
            ws1.cell(row=row_idx, column=col).alignment = center_align
    
    # Điều chỉnh độ rộng cột
    ws1.column_dimensions['A'].width = 20
    ws1.column_dimensions['B'].width = 15
    ws1.column_dimensions['C'].width = 15
    ws1.column_dimensions['D'].width = 15
    
    # ===== SHEET 2: Chi tiết tất cả sinh viên =====
    ws2 = wb.create_sheet("Chi tiết")
    
    headers2 = [
        "MSSV", "Họ tên", "Điểm TB", "Thời gian (h)", "Tham gia (%)", 
        "Nộp muộn", "Level cũ", "Level mới", "Thay đổi",
        "BT cũ", "BT mới", "Ghi chú"
    ]
    
    for col, header in enumerate(headers2, 1):
        cell = ws2.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border
    
    level_order = ['Yeu', 'Trung binh', 'Kha', 'Xuat sac']
    row_idx = 2
    
    for student_id in sorted(new_data.keys()):
        new_class = new_data[student_id]
        old_class = old_data.get(student_id, {})
        student_info = students_info.get(student_id, {})
        
        csv_data = student_info.get('csv_data', {})
        courses = student_info.get('courses', {})
        
        # Tính toán
        total_score = float(csv_data.get('total_score', 0))
        total_time = sum(float(c.get('time_minutes', 0)) for c in courses.values() if isinstance(c, dict))
        time_hours = round(total_time / 60, 1)
        attendance = round(float(csv_data.get('attendance_rate', 0)) * 100, 1)
        late_submissions = int(csv_data.get('late_submissions', 0))
        
        old_level = old_class.get('final_level', 'N/A')
        new_level = new_class.get('final_level', '')
        old_anomaly = old_class.get('anomaly_detected', False)
        new_anomaly = new_class.get('anomaly_detected', False)
        
        # Xác định thay đổi
        if old_level == 'N/A':
            change = "Mới"
            change_type = 'same'
        elif old_level == new_level:
            change = "Không đổi"
            change_type = 'same'
        else:
            old_idx = level_order.index(old_level) if old_level in level_order else -1
            new_idx = level_order.index(new_level) if new_level in level_order else -1
            if new_idx > old_idx:
                change = "↑ Tăng"
                change_type = 'up'
            else:
                change = "↓ Giảm"
                change_type = 'down'
        
        # Ghi chú
        notes = []
        if old_anomaly and not new_anomaly:
            notes.append("Bỏ cảnh báo BT")
        elif not old_anomaly and new_anomaly:
            notes.append("Thêm cảnh báo BT")
        
        # Ghi dữ liệu
        row_data = [
            student_id,
            student_info.get('name', ''),
            total_score,
            time_hours,
            attendance,
            late_submissions,
            old_level,
            new_level,
            change,
            "Có" if old_anomaly else "Không",
            "Có" if new_anomaly else "Không",
            ", ".join(notes)
        ]
        
        for col, value in enumerate(row_data, 1):
            cell = ws2.cell(row=row_idx, column=col, value=value)
            cell.border = thin_border
            cell.alignment = center_align
            
            # Tô màu theo level
            if col == 7 and old_level in level_fills:
                cell.fill = level_fills[old_level]
            elif col == 8 and new_level in level_fills:
                cell.fill = level_fills[new_level]
            elif col == 9:
                cell.fill = change_fills[change_type]
        
        row_idx += 1
    
    # Điều chỉnh độ rộng cột
    col_widths = [12, 25, 10, 12, 12, 10, 12, 12, 12, 10, 10, 25]
    for i, width in enumerate(col_widths, 1):
        ws2.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width
    
    # ===== SHEET 3: Chỉ những thay đổi =====
    ws3 = wb.create_sheet("Thay đổi")
    
    for col, header in enumerate(headers2, 1):
        cell = ws3.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border
    
    row_idx = 2
    for student_id in sorted(new_data.keys()):
        new_class = new_data[student_id]
        old_class = old_data.get(student_id, {})
        
        old_level = old_class.get('final_level', 'N/A')
        new_level = new_class.get('final_level', '')
        old_anomaly = old_class.get('anomaly_detected', False)
        new_anomaly = new_class.get('anomaly_detected', False)
        
        # Chỉ lấy những thay đổi
        if old_level == new_level and old_anomaly == new_anomaly:
            continue
        
        student_info = students_info.get(student_id, {})
        csv_data = student_info.get('csv_data', {})
        courses = student_info.get('courses', {})
        
        total_score = float(csv_data.get('total_score', 0))
        total_time = sum(float(c.get('time_minutes', 0)) for c in courses.values() if isinstance(c, dict))
        time_hours = round(total_time / 60, 1)
        attendance = round(float(csv_data.get('attendance_rate', 0)) * 100, 1)
        late_submissions = int(csv_data.get('late_submissions', 0))
        
        if old_level == 'N/A':
            change = "Mới"
            change_type = 'same'
        elif old_level == new_level:
            change = "Không đổi"
            change_type = 'same'
        else:
            old_idx = level_order.index(old_level) if old_level in level_order else -1
            new_idx = level_order.index(new_level) if new_level in level_order else -1
            if new_idx > old_idx:
                change = "↑ Tăng"
                change_type = 'up'
            else:
                change = "↓ Giảm"
                change_type = 'down'
        
        notes = []
        if old_anomaly and not new_anomaly:
            notes.append("Bỏ cảnh báo BT")
        elif not old_anomaly and new_anomaly:
            notes.append("Thêm cảnh báo BT")
        
        row_data = [
            student_id,
            student_info.get('name', ''),
            total_score,
            time_hours,
            attendance,
            late_submissions,
            old_level,
            new_level,
            change,
            "Có" if old_anomaly else "Không",
            "Có" if new_anomaly else "Không",
            ", ".join(notes)
        ]
        
        for col, value in enumerate(row_data, 1):
            cell = ws3.cell(row=row_idx, column=col, value=value)
            cell.border = thin_border
            cell.alignment = center_align
            
            if col == 7 and old_level in level_fills:
                cell.fill = level_fills[old_level]
            elif col == 8 and new_level in level_fills:
                cell.fill = level_fills[new_level]
            elif col == 9:
                cell.fill = change_fills[change_type]
        
        row_idx += 1
    
    for i, width in enumerate(col_widths, 1):
        ws3.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width
    
    # Lưu file
    wb.save(output_file)
    print(f"✅ Đã xuất file: {output_file}")

def main():
    print("=" * 80)
    print("📊 XUẤT KẾT QUẢ SO SÁNH RA EXCEL")
    print("=" * 80)
    
    # Kết nối Supabase
    supabase = create_client(SUPABASE_URL, SUPABASE_KEY)
    
    # Load dữ liệu sinh viên
    print("\n📊 Đang tải dữ liệu sinh viên...")
    generator = StudentDataGenerator(
        seed=42, 
        csv_path='student_classification_supabase_ready_final.csv',
        use_supabase=False
    )
    students = generator.load_all_students()
    students_info = {s['student_id']: s for s in students}
    print(f"   ✅ Đã tải {len(students)} sinh viên")
    
    # Load lịch sử (kết quả cũ)
    print("\n📦 Đang tải kết quả cũ từ lịch sử...")
    history = load_history()
    
    old_data = {}
    if history['versions']:
        # Lấy version cuối cùng trong lịch sử
        last_version = history['versions'][-1]
        old_data = {c['student_id']: c for c in last_version['classifications']}
        print(f"   ✅ Đã tải {len(old_data)} bản ghi cũ (version {last_version['version']})")
    else:
        print("   ⚠️ Không có lịch sử cũ")
    
    # Lấy kết quả mới từ Supabase
    print("\n📥 Đang tải kết quả mới từ Supabase...")
    new_data = get_current_classifications(supabase)
    print(f"   ✅ Đã tải {len(new_data)} bản ghi mới")
    
    # Xuất Excel
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_file = f"comparison_result_{timestamp}.xlsx"
    
    print(f"\n📝 Đang tạo file Excel...")
    create_excel_report(students_info, old_data, new_data, output_file)
    
    print("\n✅ HOÀN THÀNH!")

if __name__ == "__main__":
    main()
